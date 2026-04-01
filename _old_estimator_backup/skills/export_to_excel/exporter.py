#!/usr/bin/env python3
"""
Export to Excel — OpenClaw Estimator Department
=================================================
ФИНАЛЬНЫЙ ШАГ в SOP Chief Estimator. Генерация отчёта.

Конвертирует takeoff JSON в Excel (.xlsx) для загрузки в
сметные программы (AccuBid, Procore, Excel-based estimating).

2 ЛИСТА:
  Sheet 1: "Takeoff Summary" — все 8 секций с subtotals
  Sheet 2: "Zone Distribution" — устройства по зонам/комнатам

Форматирование: цветные заголовки, рамки, subtotals, grand total.
FREE (нет API). Только Python + openpyxl.

ВХОД: JSON из stdout любого навыка (или файл).
ВЫХОД: .xlsx файл + JSON статус в stdout.

Usage:
    python exporter.py --input takeoff_result.json --output report.xlsx
    echo '{"devices":[...]}' | python exporter.py --input - --output report.xlsx
"""

import sys
import json
import argparse
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print(json.dumps({"status": "error", "message": "openpyxl not installed. Run: pip install openpyxl"}))
    sys.exit(1)


# ==================================================================
# Константы форматирования
# ==================================================================
# SECTION_NAMES: 8 секций электрооборудования.
# Этот список = финальный. Добавлять 9-ю секцию только
# после обновления DeviceEntry Pydantic schema ВО ВСЕХ навыках!
# ==================================================================

SECTION_NAMES = {
    "1_distribution": "1. Distribution & Power Equipment",
    "2_hvac_connections": "2. HVAC & Plumbing Connections",
    "3_lighting": "3. Lighting Fixtures",
    "4_lighting_controls": "4. Lighting Controls",
    "5_receptacles": "5. Power Receptacles",
    "6_low_voltage": "6. Low Voltage (IT, AV, Security)",
    "7_fire_alarm": "7. Fire Alarm",
    "8_rough_in": "8. Rough-In Materials"
}

# Стили Excel (Calibri, синие заголовки — под AccuBid/констр.сметы)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=11)
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def create_summary_sheet(ws, devices):
    """Sheet 1: Full takeoff summary by section."""
    headers = ["Section", "Device Type", "Symbol", "Qty", "Zone", "Source", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Group devices by section
    sections = {}
    for d in devices:
        sec = d.get("section", "8_rough_in")
        sections.setdefault(sec, []).append(d)

    row = 2
    grand_total = 0

    for sec_key in sorted(SECTION_NAMES.keys()):
        sec_devices = sections.get(sec_key, [])
        if not sec_devices:
            continue

        # Section header row
        sec_name = SECTION_NAMES[sec_key]
        subtotal = sum(d.get("quantity", 0) for d in sec_devices)
        grand_total += subtotal

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=f"{sec_name} (Subtotal: {subtotal})")
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

        for d in sec_devices:
            ws.cell(row=row, column=1, value=sec_name.split(".")[0]).font = BODY_FONT
            ws.cell(row=row, column=2, value=d.get("device_type", "")).font = BODY_FONT
            ws.cell(row=row, column=3, value=d.get("symbol_on_drawing", "")).font = BODY_FONT
            ws.cell(row=row, column=4, value=d.get("quantity", 0)).font = BODY_FONT
            ws.cell(row=row, column=5, value=d.get("zone", "")).font = BODY_FONT
            ws.cell(row=row, column=6, value=d.get("notes", "")).font = BODY_FONT
            ws.cell(row=row, column=7, value=d.get("confidence", "")).font = BODY_FONT
            for c in range(1, 8):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

    # Grand total
    row += 1
    cell = ws.cell(row=row, column=1, value="GRAND TOTAL")
    cell.font = Font(name="Calibri", bold=True, size=12)
    ws.cell(row=row, column=4, value=grand_total).font = Font(name="Calibri", bold=True, size=12)

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 12


def create_zone_sheet(ws, devices):
    """Sheet 2: Device distribution by zone."""
    headers = ["Zone", "Device Type", "Qty", "Section"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Group by zone
    zones = {}
    for d in devices:
        zone = d.get("zone", "Unknown")
        zones.setdefault(zone, []).append(d)

    row = 2
    for zone_name in sorted(zones.keys()):
        zone_devs = zones[zone_name]
        for d in zone_devs:
            ws.cell(row=row, column=1, value=zone_name).font = BODY_FONT
            ws.cell(row=row, column=2, value=d.get("device_type", "")).font = BODY_FONT
            ws.cell(row=row, column=3, value=d.get("quantity", 0)).font = BODY_FONT
            ws.cell(row=row, column=4, value=SECTION_NAMES.get(d.get("section", ""), "")).font = BODY_FONT
            for c in range(1, 5):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 35


def export(devices, output_path):
    """Create Excel workbook with 2 sheets."""
    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Takeoff Summary"
    create_summary_sheet(ws1, devices)

    # Sheet 2: Zones
    ws2 = wb.create_sheet("Zone Distribution")
    create_zone_sheet(ws2, devices)

    wb.save(output_path)
    return output_path


def run(input_path, output_path):
    """Load JSON, export to Excel."""
    if input_path == "-":
        data = json.load(sys.stdin)
    else:
        if not os.path.exists(input_path):
            return {"status": "error", "message": f"File not found: {input_path}"}
        with open(input_path) as f:
            data = json.load(f)

    devices = data.get("devices", data if isinstance(data, list) else [])
    if not devices:
        return {"status": "error", "message": "No devices found in input"}

    path = export(devices, output_path)
    total = sum(d.get("quantity", 0) for d in devices)

    return {
        "status": "success",
        "output_path": os.path.abspath(path),
        "total_devices": len(devices),
        "total_quantity": total
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export Takeoff to Excel")
    parser.add_argument("--input", default="-", help="JSON input file (- for stdin)")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    result = run(args.input, args.output)
    print(json.dumps(result, ensure_ascii=False, indent=2))
